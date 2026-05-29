from __future__ import annotations
import argparse, csv, json, os, re, sqlite3
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
NL = chr(10)
TAB = chr(9)
def now_iso():
 return datetime.now(timezone.utc).isoformat(timespec='seconds')
@dataclass
class DriveFile:
 id: str
 name: str
 mime_type: str
 modified_time: str = ''
 local_path: str = ''
@dataclass
class PropertyRecord:
 source_file_id: str
 source_file_name: str
 property_name: str = '未抽出'
 address: str = ''
 price: int | None = None
 land_area: float | None = None
 building_area: float | None = None
 structure: str = ''
 built_year: str = ''
 station: str = ''
 walk_minutes: int | None = None
 gross_yield: float | None = None
 annual_rent: int | None = None
 zoning: str = ''
 road_access: str = ''
 legal_status: str = ''
 loan_score: int = 0
 full_loan_possibility: str = '要確認'
 recommended_banks: str = ''
 offer_price: int | None = None
 analysis_summary: str = ''
 risk_summary: str = ''
 raw_text_excerpt: str = ''
 created_at: str = field(default_factory=now_iso)
 updated_at: str = field(default_factory=now_iso)
 def to_dict(self) -> dict[str, Any]:
  return asdict(self)
@dataclass
class AnalysisRun:
 run_type: str
 status: str
 files_scanned: int
 files_analyzed: int
 new_properties: int
 summary: str
 created_at: str = field(default_factory=now_iso)
 def to_dict(self) -> dict[str, Any]:
  return asdict(self)
def normalize_number(text: str) -> int | None:
 s = str(text).replace(',', '').replace('，', '').replace(' ', '')
 m = re.search(r'([0-9]+(?:\.[0-9]+)?)\s*(億|万)?', s)
 if not m:
  return None
 value = float(m.group(1)); unit = m.group(2)
 if unit == '億':
  return int(value * 100000000)
 if unit == '万':
  return int(value * 10000)
 return int(value)
def first_match(patterns, text):
 for pattern in patterns:
  m = re.search(pattern, text, flags=re.MULTILINE | re.IGNORECASE)
  if m:
   return m.group(1).strip().strip('｜|,，:：').strip()
 return ''
def parse_float(patterns, text):
 raw = first_match(patterns, text)
 if not raw:
  return None
 m = re.search(r'[0-9,]+(?:\.[0-9]+)?', raw)
 return float(m.group(0).replace(',', '')) if m else None
def parse_int(patterns, text):
 raw = first_match(patterns, text)
 return normalize_number(raw) if raw else None
def analyze_property(file: DriveFile, raw_text: str) -> PropertyRecord:
 text = raw_text.replace('　', ' ')
 sep = r'[\s\t,，:：|｜]+'
 name = first_match([rf'(?:物件名|名称|建物名){sep}([^\n\r]+)', r'^#\s*([^\n\r]+)$'], text) or Path(file.name).stem
 address = first_match([rf'(?:所在地|住所|地番){sep}([^\n\r]+)', r'((?:大阪|京都|兵庫|奈良|滋賀|和歌山|東京|神奈川|千葉|埼玉).{4,60})'], text)
 price = parse_int([rf'(?:価格|売買価格|販売価格){sep}([0-9,\.]+\s*(?:億|万)?)'], text)
 rent = parse_int([rf'(?:年間賃料|満室想定賃料|年収|年間収入){sep}([0-9,\.]+\s*(?:億|万)?)'], text)
 gy = parse_float([rf'(?:利回り|表面利回り|想定利回り){sep}([0-9\.]+\s*%)'], text)
 if gy is None and price and rent:
  gy = round(rent / price * 100, 2)
 land = parse_float([rf'(?:土地面積|敷地面積){sep}([0-9,\.]+\s*(?:㎡|m2|平米)?)'], text)
 building = parse_float([rf'(?:建物面積|延床面積|専有面積){sep}([0-9,\.]+\s*(?:㎡|m2|平米)?)'], text)
 walk_raw = first_match([r'徒歩\s*([0-9]+)\s*分', rf'駅徒歩{sep}([0-9]+)\s*分'], text)
 walk = int(walk_raw) if walk_raw.isdigit() else None
 structure = first_match([rf'(?:構造){sep}([^\n\r]+)', r'(RC造|鉄筋コンクリート造|S造|鉄骨造|木造|軽量鉄骨造)'], text)
 built = first_match([rf'(?:築年|築年月|竣工|建築年月){sep}([^\n\r]+)', r'(19[0-9]{2}年|20[0-9]{2}年|築\s*[0-9]+\s*年)'], text)
 station = first_match([rf'(?:最寄駅|交通){sep}([^\n\r]+)', r'([^\n\r]+駅\s*徒歩\s*[0-9]+\s*分)'], text)
 zoning = first_match([rf'(?:用途地域){sep}([^\n\r]+)'], text)
 road = first_match([rf'(?:接道|道路){sep}([^\n\r]+)'], text)
 legal = first_match([rf'(?:権利|土地権利){sep}([^\n\r]+)'], text)
 score = score_property(price, gy, structure, walk, land, text)
 poss = '高' if score >= 75 else '中' if score >= 55 else '低〜要確認'
 banks = recommend_banks(address, structure, score)
 offer = int(price * 0.92) if price else None
 summary = build_summary(name, price, gy, poss, score)
 risk = build_risk_summary(text, structure, built)
 return PropertyRecord(file.id, file.name, name[:120], address[:200], price, land, building, structure[:80], built[:80], station[:120], walk, gy, rent, zoning[:120], road[:120], legal[:120], score, poss, banks, offer, summary, risk, text[:1000])
def score_property(price, gy, structure, walk, land, text):
 score = 40
 if gy is not None:
  score += 25 if gy >= 9 else 18 if gy >= 8 else 10 if gy >= 7 else -5
 if structure and any(k in structure for k in ['RC', '鉄筋', '鉄骨', 'S造']):
  score += 10
 elif '木造' in structure:
  score += 4
 if walk is not None:
  score += 10 if walk <= 10 else 3 if walk <= 15 else -5
 if land and land >= 100:
  score += 6
 if price and price <= 100000000:
  score += 4
 score -= sum(8 for word in ['再建築不可', '既存不適格', '違法', '私道', 'セットバック', '借地', '告知事項'] if word in text)
 return max(0, min(100, score))
def recommend_banks(address, structure, score):
 base = ['大阪信用金庫', '大阪厚生信用金庫', '大阪商工信用金庫', '近畿産業信用組合', '関西みらい銀行']
 if '京都' in address:
  base.insert(0, '京都中央信用金庫')
 if '兵庫' in address or '神戸' in address:
  base.insert(0, '兵庫信用金庫')
 if score < 55:
  base.append('日本政策金融公庫（小口・修繕資金の相談）')
 if 'RC' in structure or '鉄筋' in structure:
  base.append('地銀・信金の長期融資枠')
 return '、'.join(dict.fromkeys(base[:6]))
def build_summary(name, price, gy, poss, score):
 pt = f'{price:,}円' if price else '価格未抽出'
 yt = f'{gy:.2f}%' if gy is not None else '利回り未抽出'
 return f'{name} は {pt} / 表面利回り {yt}。融資スコアは {score}/100、フルローン可能性は {poss}。'
def build_risk_summary(text, structure, built):
 found = [w for w in ['再建築不可', '既存不適格', '違法', '私道', 'セットバック', '借地', '告知事項', 'ハザード'] if w in text]
 notes = []
 if found:
  notes.append('要確認キーワード: ' + '、'.join(found))
 if not structure:
  notes.append('構造が未抽出のため耐用年数確認が必要')
 if not built:
  notes.append('築年が未抽出のため融資期間確認が必要')
 return '。'.join(notes) if notes else '大きなリスクキーワードは自動検出されていません。接道・法令・修繕履歴は別途確認してください。'
def extract_text(path: Path):
 suffix = path.suffix.lower()
 if suffix == '.csv':
  return NL.join(TAB.join(row) for row in csv.reader(path.open(encoding='utf-8-sig', errors='ignore')))
 if suffix in ['.txt', '.md']:
  return path.read_text(encoding='utf-8', errors='ignore')
 if suffix == '.pdf':
  from pypdf import PdfReader
  return NL.join(page.extract_text() or '' for page in PdfReader(str(path)).pages)
 if suffix in ['.xlsx', '.xlsm']:
  from openpyxl import load_workbook
  wb = load_workbook(path, read_only=True, data_only=True); lines = []
  for ws in wb.worksheets:
   for row in ws.iter_rows(values_only=True):
    vals = [str(v) for v in row if v not in (None, '')]
    if vals:
     lines.append(TAB.join(vals))
  return NL.join(lines)
 return path.read_text(encoding='utf-8', errors='ignore')
def list_drive_files(folder_id: str):
 raw = os.environ.get('GOOGLE_SERVICE_ACCOUNT_JSON', '').strip()
 if not raw:
  raise RuntimeError('GOOGLE_SERVICE_ACCOUNT_JSON is not set. See docs/setup.md.')
 from google.oauth2 import service_account
 from googleapiclient.discovery import build
 info = json.loads(raw)
 creds = service_account.Credentials.from_service_account_info(info, scopes=['https://www.googleapis.com/auth/drive.readonly'])
 service = build('drive', 'v3', credentials=creds, cache_discovery=False)
 q = f"'{folder_id}' in parents and trashed = false"
 res = service.files().list(q=q, fields='files(id,name,mimeType,modifiedTime)', supportsAllDrives=True, includeItemsFromAllDrives=True, pageSize=100).execute()
 return [DriveFile(x['id'], x['name'], x.get('mimeType', ''), x.get('modifiedTime', '')) for x in res.get('files', [])]
def download_drive_file(file, download_dir: Path):
 from google.oauth2 import service_account
 from googleapiclient.discovery import build
 from googleapiclient.http import MediaIoBaseDownload
 import io
 info = json.loads(os.environ['GOOGLE_SERVICE_ACCOUNT_JSON'])
 creds = service_account.Credentials.from_service_account_info(info, scopes=['https://www.googleapis.com/auth/drive.readonly'])
 service = build('drive', 'v3', credentials=creds, cache_discovery=False)
 download_dir.mkdir(parents=True, exist_ok=True)
 safe = ''.join(c if c.isalnum() or c in '._- ' else '_' for c in file.name)
 path = download_dir / f'{file.id}_{safe}'
 req = service.files().get_media(fileId=file.id, supportsAllDrives=True)
 with io.FileIO(path, 'wb') as fh:
  dl = MediaIoBaseDownload(fh, req); done = False
  while not done:
   _, done = dl.next_chunk()
 file.local_path = str(path); return file
FIELDS = ['property_name','address','price','gross_yield','annual_rent','loan_score','full_loan_possibility','recommended_banks','offer_price','land_area','building_area','structure','built_year','station','walk_minutes','zoning','road_access','legal_status','analysis_summary','risk_summary','source_file_name','source_file_id']
DB_COLUMNS = ['source_file_id','source_file_name','property_name','address','price','land_area','building_area','structure','built_year','station','walk_minutes','gross_yield','annual_rent','zoning','road_access','legal_status','loan_score','full_loan_possibility','recommended_banks','offer_price','analysis_summary','risk_summary','raw_text_excerpt','created_at','updated_at']
def init_db(db_path: Path):
 db_path.parent.mkdir(parents=True, exist_ok=True)
 con = sqlite3.connect(db_path)
 con.execute('pragma journal_mode=wal')
 con.execute('create table if not exists drive_files (id text primary key, name text, mime_type text, modified_time text, local_path text, created_at text, updated_at text)')
 con.execute('create table if not exists properties (id integer primary key autoincrement, source_file_id text unique, source_file_name text, property_name text, address text, price integer, land_area real, building_area real, structure text, built_year text, station text, walk_minutes integer, gross_yield real, annual_rent integer, zoning text, road_access text, legal_status text, loan_score integer, full_loan_possibility text, recommended_banks text, offer_price integer, analysis_summary text, risk_summary text, raw_text_excerpt text, created_at text, updated_at text)')
 con.execute('create table if not exists analysis_runs (id integer primary key autoincrement, run_type text, status text, files_scanned integer, files_analyzed integer, new_properties integer, summary text, created_at text)')
 con.execute('create index if not exists idx_properties_score on properties(loan_score desc)')
 con.execute('create index if not exists idx_properties_price on properties(price)')
 con.execute('create index if not exists idx_properties_address on properties(address)')
 con.execute('create index if not exists idx_runs_created on analysis_runs(created_at desc)')
 return con
def save_database(records, run, output_dir: Path):
 db_path = output_dir / 'property_ocr.db'
 con = init_db(db_path)
 con.execute('insert into analysis_runs (run_type,status,files_scanned,files_analyzed,new_properties,summary,created_at) values (?,?,?,?,?,?,?)', (run.run_type, run.status, run.files_scanned, run.files_analyzed, run.new_properties, run.summary, run.created_at))
 placeholders = ','.join(['?'] * len(DB_COLUMNS))
 updates = ','.join([c + '=excluded.' + c for c in DB_COLUMNS if c != 'source_file_id'])
 columns = ','.join(DB_COLUMNS)
 sql = 'insert into properties (' + columns + ') values (' + placeholders + ') on conflict(source_file_id) do update set ' + updates
 for rec in records:
  d = rec.to_dict()
  con.execute(sql, tuple(d.get(c) for c in DB_COLUMNS))
  con.execute('insert or replace into drive_files (id,name,mime_type,modified_time,local_path,created_at,updated_at) values (?,?,?,?,?,?,?)', (rec.source_file_id, rec.source_file_name, '', '', '', rec.created_at, rec.updated_at))
 con.commit(); con.close()
 write_query_examples(output_dir / 'query_examples.sql')
 return db_path
def write_query_examples(path: Path):
 queries = [
  '-- 融資スコア上位10件',
  'select property_name,address,price,gross_yield,loan_score,full_loan_possibility,recommended_banks from properties order by loan_score desc limit 10;',
  '-- 大阪で利回り8%以上',
  "select property_name,address,price,gross_yield from properties where address like '%大阪%' and gross_yield >= 8 order by gross_yield desc;",
  '-- 指値候補',
  'select property_name,price,offer_price,loan_score,risk_summary from properties where offer_price is not null order by loan_score desc;',
  '-- 最新実行履歴',
  'select * from analysis_runs order by created_at desc limit 5;',
 ]
 path.write_text(NL.join(queries) + NL, encoding='utf-8')

def write_reports(records, run, output_dir: Path):
 output_dir.mkdir(parents=True, exist_ok=True)
 rows = sorted([r.to_dict() for r in records], key=lambda r: r.get('loan_score') or 0, reverse=True)
 with (output_dir / 'property_report.csv').open('w', encoding='utf-8-sig', newline='') as f:
  w = csv.DictWriter(f, fieldnames=FIELDS, extrasaction='ignore'); w.writeheader(); w.writerows(rows)
 from openpyxl import Workbook
 wb = Workbook(); ws = wb.active; ws.title = 'property_report'; ws.append(FIELDS)
 for row in rows:
  ws.append([row.get(k) for k in FIELDS])
 wb.save(output_dir / 'property_report.xlsx')
 lines = ['# 物件OCR・分析レポート', '', f'実行状態: {run.status}', f'スキャン数: {run.files_scanned}', f'分析数: {run.files_analyzed}', '', '## 融資スコア順ランキング', '']
 for i, row in enumerate(rows, 1):
  lines += [f"### {i}. {row.get('property_name')}", f"- 所在地: {row.get('address') or '未抽出'}", f"- 価格: {row.get('price') or '未抽出'}", f"- 表面利回り: {row.get('gross_yield') or '未抽出'}%", f"- 融資スコア: {row.get('loan_score')}/100", f"- フルローン可能性: {row.get('full_loan_possibility')}", f"- 想定銀行: {row.get('recommended_banks')}", f"- コメント: {row.get('analysis_summary')}", f"- リスク: {row.get('risk_summary')}", '']
 (output_dir / 'property_report.txt').write_text(NL.join(lines), encoding='utf-8')
 (output_dir / 'property_report.json').write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding='utf-8')
 (output_dir / 'analysis_run.json').write_text(json.dumps(run.to_dict(), ensure_ascii=False, indent=2), encoding='utf-8')
 save_database(records, run, output_dir)
def run_pipeline(folder_id: str, output_dir: Path, download_dir: Path, input_dir: Path | None = None):
 files = []; local = []
 if input_dir and input_dir.exists():
  for p in sorted(input_dir.iterdir()):
   if p.is_file() and not p.name.startswith('.'):
    df = DriveFile(p.stem, p.name, 'local', local_path=str(p)); files.append(df); local.append((df, p))
 else:
  files = list_drive_files(folder_id)
  for df in files:
   dl = download_drive_file(df, download_dir); local.append((dl, Path(dl.local_path)))
 records = []
 errors = []
 for df, p in local:
  try:
   records.append(analyze_property(df, extract_text(p)))
  except Exception as exc:
   errors.append(f'{df.name}: {exc}')
 status = 'success' if not errors else 'partial_success'
 summary = f'Analyzed {len(records)} of {len(files)} files.'
 if errors:
  summary = summary + ' Errors: ' + ' | '.join(errors)
 run = AnalysisRun('github-actions' if os.environ.get('GITHUB_ACTIONS') else 'local', status, len(files), len(records), len(records), summary)
 write_reports(records, run, output_dir)
 print(run.summary)
 return 0 if records or not files else 2
def query_db(db_path: Path, sql: str):
 con = sqlite3.connect(db_path)
 con.row_factory = sqlite3.Row
 rows = [dict(r) for r in con.execute(sql).fetchall()]
 con.close()
 print(json.dumps(rows, ensure_ascii=False, indent=2))
 return 0
def main(argv=None):
 parser = argparse.ArgumentParser(description='Property OCR pipeline with SQLite output')
 sub = parser.add_subparsers(dest='command', required=True)
 r = sub.add_parser('run'); r.add_argument('--folder-id', default=os.environ.get('TARGET_DRIVE_FOLDER_ID', '11cA-CrY7rjlQlzdXywpT3i7PLRrXxOgD')); r.add_argument('--input-dir', default=''); r.add_argument('--output-dir', default='outputs'); r.add_argument('--download-dir', default='downloads')
 q = sub.add_parser('query'); q.add_argument('--db', default='outputs/property_ocr.db'); q.add_argument('--sql', required=True)
 args = parser.parse_args(argv)
 if args.command == 'query':
  return query_db(Path(args.db), args.sql)
 return run_pipeline(args.folder_id, Path(args.output_dir), Path(args.download_dir), Path(args.input_dir) if args.input_dir else None)


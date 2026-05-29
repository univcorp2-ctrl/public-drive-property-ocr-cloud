from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return extract_pdf_text(path)
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return extract_xlsx_text(path)
    if suffix == ".csv":
        return extract_csv_text(path)
    if suffix in {".txt", ".md"}:
        return path.read_text(encoding="utf-8", errors="ignore")
    if suffix in {".png", ".jpg", ".jpeg", ".webp", ".tif", ".tiff"}:
        return extract_image_text(path)
    return path.read_text(encoding="utf-8", errors="ignore")


def extract_pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    pages = []
    for page in reader.pages:
        pages.append(page.extract_text() or "")
    return "\n".join(pages).strip()


def extract_xlsx_text(path: Path) -> str:
    wb = load_workbook(path, read_only=True, data_only=True)
    lines: list[str] = []
    for ws in wb.worksheets:
        lines.append(f"# sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            values = [str(v) for v in row if v not in (None, "")]
            if values:
                lines.append("\t".join(values))
    return "\n".join(lines)


def extract_csv_text(path: Path) -> str:
    lines: list[str] = []
    with path.open("r", encoding="utf-8-sig", errors="ignore", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            lines.append("\t".join(row))
    return "\n".join(lines)


def extract_image_text(path: Path) -> str:
    try:
        import pytesseract  # type: ignore
        from PIL import Image
    except Exception:
        return f"[画像OCR未実行] {path.name}: pytesseract が未インストールです。必要なら runner に Tesseract を追加してください。"
    return pytesseract.image_to_string(Image.open(path), lang="jpn+eng")

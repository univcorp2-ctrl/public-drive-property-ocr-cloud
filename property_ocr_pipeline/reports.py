from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import AnalysisRun, PropertyRecord

FIELDS = [
    "property_name",
    "address",
    "price",
    "gross_yield",
    "annual_rent",
    "loan_score",
    "full_loan_possibility",
    "recommended_banks",
    "offer_price",
    "land_area",
    "building_area",
    "structure",
    "built_year",
    "station",
    "walk_minutes",
    "zoning",
    "road_access",
    "legal_status",
    "analysis_summary",
    "risk_summary",
    "source_file_name",
    "source_file_id",
]


def write_reports(records: Iterable[PropertyRecord], run: AnalysisRun, output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    rows = sorted([r.to_dict() for r in records], key=lambda r: r.get("loan_score") or 0, reverse=True)
    write_csv(rows, output_dir / "property_report.csv")
    write_xlsx(rows, output_dir / "property_report.xlsx")
    write_txt(rows, run, output_dir / "property_report.txt")
    (output_dir / "property_report.json").write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    (output_dir / "analysis_run.json").write_text(json.dumps(run.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(rows: list[dict], path: Path) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "property_report"
    ws.append(FIELDS)
    for row in rows:
        ws.append([row.get(field) for field in FIELDS])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for col_idx, field in enumerate(FIELDS, start=1):
        max_len = max([len(str(field))] + [len(str(r.get(field, ""))) for r in rows])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def write_txt(rows: list[dict], run: AnalysisRun, path: Path) -> None:
    lines = [
        "# 物件OCR・分析レポート",
        "",
        f"実行状態: {run.status}",
        f"スキャン数: {run.files_scanned}",
        f"分析数: {run.files_analyzed}",
        "",
        "## 融資スコア順ランキング",
        "",
    ]
    for idx, row in enumerate(rows, start=1):
        lines.extend(
            [
                f"### {idx}. {row.get('property_name')}",
                f"- 所在地: {row.get('address') or '未抽出'}",
                f"- 価格: {format_yen(row.get('price'))}",
                f"- 表面利回り: {row.get('gross_yield') or '未抽出'}%",
                f"- 融資スコア: {row.get('loan_score')}/100",
                f"- フルローン可能性: {row.get('full_loan_possibility')}",
                f"- 想定銀行: {row.get('recommended_banks')}",
                f"- 指値目安: {format_yen(row.get('offer_price'))}",
                f"- コメント: {row.get('analysis_summary')}",
                f"- リスク: {row.get('risk_summary')}",
                "",
            ]
        )
    path.write_text("\n".join(lines), encoding="utf-8")


def format_yen(value: object) -> str:
    if isinstance(value, int):
        return f"{value:,}円"
    return "未抽出"

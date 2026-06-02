import csv
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def _flatten(prefix: str, value: Any, rows: list[tuple[str, Any]]) -> None:
    if isinstance(value, dict):
        for key, child in value.items():
            _flatten(f"{prefix}.{key}" if prefix else key, child, rows)
    elif isinstance(value, list):
        rows.append((prefix, " / ".join(str(item) for item in value)))
    else:
        rows.append((prefix, value))


def export_result(job_id: str, output_dir: Path, result: dict[str, Any], extracted_text: str) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    rows: list[tuple[str, Any]] = []
    _flatten("", result, rows)

    csv_path = output_dir / f"{job_id}.csv"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as fp:
        writer = csv.writer(fp)
        writer.writerow(["key", "value"])
        writer.writerows(rows)

    txt_path = output_dir / f"{job_id}.txt"
    with txt_path.open("w", encoding="utf-8") as fp:
        fp.write("ボリューム検討結果\n")
        fp.write("=" * 24 + "\n\n")
        for key, value in rows:
            fp.write(f"{key}: {value}\n")
        fp.write("\n--- extracted text ---\n")
        fp.write(extracted_text)

    xlsx_path = output_dir / f"{job_id}.xlsx"
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["key", "value"])
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
    for row in rows:
        summary.append(list(row))
    summary.column_dimensions["A"].width = 40
    summary.column_dimensions["B"].width = 60

    text_sheet = workbook.create_sheet("ExtractedText")
    text_sheet.append(["text"])
    text_sheet["A1"].font = Font(bold=True)
    text_sheet.append([extracted_text])
    text_sheet.column_dimensions["A"].width = 100
    workbook.save(xlsx_path)

    return {"csv": csv_path, "txt": txt_path, "xlsx": xlsx_path}

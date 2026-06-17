from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from property_ocr.extract import OcrRecord

FIELDNAMES = [
    "source_file",
    "file_type",
    "property_name",
    "address",
    "price",
    "land_area",
    "building_area",
    "layout",
    "built_date",
    "transport",
    "phone",
    "url",
    "warnings",
    "text",
]


def _safe_sheet_value(value: str) -> str:
    # Excel cells are limited to 32,767 chars. Keep a useful prefix and avoid write errors.
    if len(value) > 32700:
        return value[:32700] + "\n...[truncated]"
    return value


def write_csv(records: list[OcrRecord], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=FIELDNAMES)
        writer.writeheader()
        for record in records:
            row = record.to_row()
            writer.writerow({field: row.get(field, "") for field in FIELDNAMES})


def write_excel(records: list[OcrRecord], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "properties"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    sheet.append(FIELDNAMES)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for record in records:
        row = record.to_row()
        sheet.append([_safe_sheet_value(str(row.get(field, ""))) for field in FIELDNAMES])

    sheet.freeze_panes = "A2"
    for column_index, field in enumerate(FIELDNAMES, start=1):
        width = 18
        if field in {"source_file", "address", "url", "warnings"}:
            width = 36
        if field == "text":
            width = 80
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path)


def write_text_files(records: list[OcrRecord], directory: Path) -> None:
    directory.mkdir(parents=True, exist_ok=True)
    all_parts: list[str] = []
    for index, record in enumerate(records, start=1):
        source_name = Path(record.source_file).name or f"record-{index}"
        safe_name = source_name.replace("/", "_").replace("\\", "_")
        text_path = directory / f"{index:03d}_{safe_name}.txt"
        body = f"SOURCE: {record.source_file}\nWARNINGS: {record.warnings}\n\n{record.text}\n"
        text_path.write_text(body, encoding="utf-8")
        all_parts.append(body)
    (directory / "all_text.txt").write_text("\n\n".join(all_parts), encoding="utf-8")


def write_summary(records: list[OcrRecord], path: Path) -> None:
    summary = {
        "record_count": len(records),
        "records_with_text": sum(1 for record in records if record.text.strip()),
        "records_with_warnings": sum(1 for record in records if record.warnings.strip()),
        "output_files": ["properties.csv", "properties.xlsx", "summary.json", "text/"],
    }
    path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")


def write_outputs(records: list[OcrRecord], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(records, output_dir / "properties.csv")
    write_excel(records, output_dir / "properties.xlsx")
    write_text_files(records, output_dir / "text")
    write_summary(records, output_dir / "summary.json")
